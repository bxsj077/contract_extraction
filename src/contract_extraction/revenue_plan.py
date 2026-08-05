from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .system_models import ContractStructured, Difference


SS = "urn:schemas-microsoft-com:office:spreadsheet"
ISO_DATE = re.compile(r"^20\d{2}-\d{1,2}-\d{1,2}$")


def _row_values(row: ET.Element) -> list[str]:
    values: list[str] = []
    column = 1
    for cell in row.findall(f"{{{SS}}}Cell"):
        index = cell.get(f"{{{SS}}}Index")
        if index:
            column = int(index)
        while len(values) < column - 1:
            values.append("")
        data = cell.find(f"{{{SS}}}Data")
        values.append("" if data is None or data.text is None else data.text.strip())
        merge = int(cell.get(f"{{{SS}}}MergeAcross", "0"))
        values.extend([""] * merge)
        column += 1 + merge
    while values and values[-1] == "":
        values.pop()
    return values


def _spreadsheetml(path: Path) -> dict[str, list[list[str]]]:
    root = ET.fromstring(path.read_text(encoding="utf-8").lstrip())
    result: dict[str, list[list[str]]] = {}
    for sheet in root.findall(f"{{{SS}}}Worksheet"):
        name = sheet.get(f"{{{SS}}}Name", "")
        table = sheet.find(f"{{{SS}}}Table")
        result[name] = [] if table is None else [_row_values(row) for row in table.findall(f"{{{SS}}}Row")]
    return result


def _xlsx(path: Path) -> dict[str, list[list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {sheet.title: [["" if value is None else str(value).strip() for value in row]
                              for row in sheet.iter_rows(values_only=True)] for sheet in workbook.worksheets}
    finally:
        workbook.close()


def read_plan_workbook(path: Path) -> dict[str, list[list[str]]]:
    prefix = path.read_bytes()[:200].lstrip()
    if prefix.startswith(b"<?xml") and b"Workbook" in prefix:
        return _spreadsheetml(path)
    if path.suffix.lower() == ".xlsx" or prefix.startswith(b"PK"):
        return _xlsx(path)
    raise ValueError("暂不支持二进制旧版XLS，请另存为XLSX或Excel 2003 XML")


def _as_date(value: str) -> date | None:
    text = value.strip()[:10]
    if not ISO_DATE.fullmatch(text):
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _find_header(rows: list[list[str]], start_label: str, end_label: str) -> tuple[int, int, int] | None:
    for row_index, row in enumerate(rows):
        if start_label in row and end_label in row:
            return row_index, row.index(start_label), row.index(end_label)
    return None


def extract_plan_periods(path: Path) -> list[dict[str, Any]]:
    periods: list[dict[str, Any]] = []
    for sheet_name, rows in read_plan_workbook(path).items():
        definitions = []
        if "收入" in sheet_name:
            definitions.append(("收入计划", "起始日期", "终止日期"))
        if "收款" in sheet_name:
            definitions.append(("收款计划", "合同履约起始日期", "合同履约截止日期"))
        for plan_type, start_label, end_label in definitions:
            found = _find_header(rows, start_label, end_label)
            if not found:
                continue
            header_row, start_col, end_col = found
            for row_number, row in enumerate(rows[header_row + 1:], header_row + 2):
                if max(start_col, end_col) >= len(row):
                    continue
                start, end = _as_date(row[start_col]), _as_date(row[end_col])
                if start and end:
                    periods.append({"plan_type": plan_type, "start_date": start.isoformat(),
                                    "end_date": end.isoformat(), "file_name": path.name,
                                    "sheet_name": sheet_name, "row_number": row_number})
    return periods


def compare_income_collection_plan(contract: ContractStructured | None, files: list[str],
                                   tolerance_days: int = 31) -> list[Difference]:
    if not files:
        return []
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    errors: list[str] = []
    for file_name in files:
        try:
            for period in extract_plan_periods(Path(file_name)):
                grouped[period["plan_type"]].append(period)
        except Exception as exc:
            errors.append(f"{Path(file_name).name}：{exc}")
    results: list[Difference] = []
    if errors:
        results.append(Difference("收入收款计划", "计划文件解析失败", "待确认", "PL-000", "收入/收款计划",
                                  "；".join(errors), needs_review=True))
    if not grouped:
        results.append(Difference("收入收款计划", "未提取到履约期间", "待确认", "PL-001", "收入/收款计划",
                                  "计划文件存在，但未提取到可比较的起止日期。", needs_review=True))
        return results
    for plan_type, periods in grouped.items():
        plan_start = min(date.fromisoformat(x["start_date"]) for x in periods)
        plan_end = max(date.fromisoformat(x["end_date"]) for x in periods)
        source = {"计划类型": plan_type, "计划起始日期": plan_start.isoformat(), "计划截止日期": plan_end.isoformat(),
                  "计划行数": len(periods), "来源文件": "；".join(sorted({x["file_name"] for x in periods})),
                  "来源工作表": "；".join(sorted({x["sheet_name"] for x in periods}))}
        if not contract or not contract.time_plan.start_date or not contract.time_plan.finish_date:
            results.append(Difference("收入收款计划", "合同履约日期缺失", "待确认", "PL-002", plan_type,
                f"{plan_type}覆盖{plan_start}至{plan_end}，但前向合同未提取到可比较的履约起止日期。",
                source, {}, needs_review=True))
            continue
        contract_start = date.fromisoformat(contract.time_plan.start_date)
        contract_end = date.fromisoformat(contract.time_plan.finish_date)
        start_gap, end_gap = (plan_start - contract_start).days, (plan_end - contract_end).days
        contract_data = {"合同履约起始日期": contract_start.isoformat(), "合同履约截止日期": contract_end.isoformat(),
                         "计划开始偏差天数": start_gap, "计划截止偏差天数": end_gap,
                         "合同工期原文": contract.time_plan.duration_raw}
        if start_gap == 0 and end_gap == 0:
            status, risk, review = "一致", "无风险", False
        elif max(abs(start_gap), abs(end_gap)) <= tolerance_days:
            status, risk, review = "基本一致，需复核偏差", "待确认", True
        else:
            status, risk, review = "明显不一致", "中风险", True
        description = (f"合同履约期{contract_start}至{contract_end}；{plan_type}{plan_start}至{plan_end}；"
                       f"计划开始相对合同{start_gap:+d}天，截止相对合同{end_gap:+d}天。")
        results.append(Difference("收入收款计划", status, risk, "PL-003", plan_type, description,
                                  source, contract_data,
                                  [f"{x['file_name']}#{x['sheet_name']}!{x['row_number']}" for x in periods], review))
    return results
