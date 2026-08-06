from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .storage import ReviewStore


HEADER_CN = {
    "project_code": "结构化项目编码", "direction": "结构化合同方向", "contract_number": "合同编号",
    "contract_name": "合同名称", "party_a": "甲方", "party_b": "乙方", "amount_yuan": "合同金额（元）",
    "sign_date": "签订日期", "effective_date": "生效日期", "contract_type": "合同性质",
    "time_plan.duration_value": "工期数值", "time_plan.duration_unit": "工期单位",
    "time_plan.start_condition_type": "起算条件类型", "time_plan.start_condition_text": "起算条件原文",
    "time_plan.start_date": "实际起算日期", "time_plan.finish_date": "预计完成日期",
    "time_plan.completion_node": "完成节点", "time_plan.fixed_deadline": "固定截止日期",
    "time_plan.duration_raw": "工期约定原文", "time_plan.calculation_status": "工期计算状态",
    "time_plan.duration_conclusion": "工期提取结论",
    "time_plan.confidence": "工期置信度", "key_clauses.服务内容": "服务内容条款",
    "key_clauses.乙方义务": "乙方义务条款", "key_clauses.关键条款": "其他关键条款",
    "parse_metadata.file_hash": "文件哈希", "parse_metadata.parse_version": "解析版本",
    "parse_metadata.aggregation": "后向汇总方式", "parse_metadata.contract_count": "后向合同数量",
    "procurement_involved": "是否涉及货物采购", "procurement_note": "货物采购说明",
    "standard_name": "标准名称", "original_name": "合同原始名称", "brand": "品牌",
    "model": "型号规格", "unit": "单位", "quantity": "数量", "technical_parameters": "技术参数",
    "evidence_id": "证据编号", "confidence": "提取置信度", "list_type": "清单性质",
    "category": "差异类别", "status": "判断结果", "risk_level": "风险等级", "rule_id": "规则编号",
    "title": "审查事项", "description": "风险说明", "forward": "前向结构化内容",
    "backward": "后向结构化内容", "evidence_ids": "证据编号", "needs_review": "是否需人工复核",
    "node": "时间节点", "difference": "差异说明", "id": "复核编号", "resolution": "复核结论",
    "created_at": "创建时间", "resolved_at": "复核完成时间",
    "contract_key": "合同标识", "field_path": "纠正字段", "corrected_value": "人工确认值",
    "note": "纠正说明", "updated_at": "更新时间",
}


def _header_cn(value: str) -> str:
    if value in HEADER_CN:
        return HEADER_CN[value]
    prefix = "time_plan.milestones."
    if value.startswith(prefix):
        return f"时间节点-{value[len(prefix):]}"
    detail_prefix = "time_plan.milestone_details."
    return f"时间节点明细-{value[len(detail_prefix):]}" if value.startswith(detail_prefix) else value


def _flatten(prefix: str, value: Any, row: dict[str, Any]) -> None:
    if isinstance(value, dict):
        for key, child in value.items():
            _flatten(f"{prefix}.{key}" if prefix else key, child, row)
    elif not isinstance(value, list):
        row[prefix] = value


def _cell_value(value: Any) -> Any:
    if isinstance(value, dict):
        if "standard_name" in value:
            parts = [f"名称：{value.get('standard_name', '')}"]
            if value.get("model"): parts.append(f"型号：{value['model']}")
            if value.get("quantity") is not None: parts.append(f"数量：{value['quantity']}{value.get('unit', '')}")
            return "；".join(parts)
        if "calculation_status" in value:
            return "；".join(x for x in (value.get("duration_raw", ""), value.get("calculation_status", ""),
                f"起算日期：{value.get('start_date')}" if value.get("start_date") else "",
                f"预计完成：{value.get('finish_date')}" if value.get("finish_date") else "") if x)
        if "计算状态" in value:
            return "；".join(x for x in (
                str(value.get("计算状态", "")),
                f"相对期限：{value.get('相对期限')}" if value.get("相对期限") else "",
                str(value.get("原文", "")),
                f"计算日期：{value.get('计算日期')}" if value.get("计算日期") else "",
            ) if x)
        if "scope_item" in value:
            return "；".join(x for x in (str(value.get("scope_item", "")), str(value.get("responsibility", "")),
                                          str(value.get("original_text", ""))) if x)
        text = json.dumps(value, ensure_ascii=False)
        return text if len(text) <= 220 else text[:217] + "..."
    if isinstance(value, list):
        text = "；".join(str(x) for x in value)
        return text if len(text) <= 220 else text[:217] + "..."
    if isinstance(value, str) and len(value) > 220:
        return value[:217] + "..."
    return value


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    rows = [{_header_cn(key): value for key, value in row.items()} for row in rows]
    headers = list(dict.fromkeys(key for row in rows for key in row)) or ["无数据"]
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(h, "")) for h in headers])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_number in range(2, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 42
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, len(h) * 2), 35)


NO_CONTRACT_RULE = "合同中无明确规定"


def _compact_clause(value: Any, max_length: int = 46) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    text = re.sub(r"^未明确[：:]\s*", "", text)
    if not text or re.match(r"^(合同未约定|合同中无明确规定|未提取)", text):
        return NO_CONTRACT_RULE
    text = re.sub(r"[（(]需查阅[^）)]*[）)]", "", text).strip()
    parts = [part.strip() for part in re.split(r"[。；;]", text) if part.strip()]
    result = parts[0] if parts else text
    return result if len(result) <= max_length else result[:max_length] + "…"


def _duration_display(plan: dict[str, Any]) -> str:
    """Return only amount+unit, or one concise non-numeric requirement."""
    value = plan.get("duration_value")
    unit = str(plan.get("duration_unit") or "").strip()
    if value not in (None, ""):
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return f"{value}{unit}"
    return _compact_clause(plan.get("duration_conclusion") or plan.get("duration_raw"))


def _milestone_requirement(raw: str, node: str, status: str) -> str:
    if re.search(r"时间异常|日期异常|早于项目起算|早于.*完工", status):
        return NO_CONTRACT_RULE
    text = re.sub(r"\s+", "", raw or "")
    relative_patterns = (
        r"((?:子)?合同(?:签订|签署|生效)(?:后|之日起)\d+(?:个?工作日|日历日|日历天|天|日|个月|月|年)(?:内)?)",
        r"((?:收到|接到).{0,12}?(?:开工令|通知)(?:后|之日起)\d+(?:个?工作日|日历日|日历天|天|日|个月|月|年)(?:内)?)",
        rf"((?:到货|交货|初验|终验|项目完工)(?:后|之日起)\d+(?:个?工作日|日历日|日历天|天|日|个月|月|年)(?:内)?)",
    )
    for pattern in relative_patterns:
        if match := re.search(pattern, text):
            return match.group(1)
    if re.search(r"(?:供货|交货)(?:结束|完成)?并(?:初步验收|初验)合格后", text):
        return "供货完成并初验合格后（未明确日期）"
    if re.search(r"(?:供货|交货)(?:结束|完成)后", text):
        return "供货完成后（未明确日期）"
    if status and not re.search(r"合同未约定|未提取", status):
        return f"合同提及{node}，未明确日期"
    return NO_CONTRACT_RULE


def _milestone_summary(plan: dict[str, Any], node: str) -> tuple[str, str]:
    detail = (plan.get("milestone_details") or {}).get(node, {}) or {}
    relative = str(detail.get("相对期限") or "").strip()
    raw_calculated = str(detail.get("计算日期") or (plan.get("milestones") or {}).get(node, "") or "").strip()
    calculated = raw_calculated if re.fullmatch(r"20\d{2}-\d{2}-\d{2}", raw_calculated) else ""
    status = str(detail.get("计算状态") or "合同未约定该节点").strip()
    if calculated:
        return calculated, calculated
    if relative:
        return _compact_clause(relative), ""
    raw = str(detail.get("原文") or "").strip()
    if raw:
        return _milestone_requirement(raw, node, status), ""
    if status and not re.search(r"合同未约定|未提取", status):
        return _milestone_requirement("", node, status), ""
    return NO_CONTRACT_RULE, ""


def _contract_time_row(project_code: str, direction: str, contract: dict[str, Any]) -> dict[str, Any]:
    plan = contract.get("time_plan", {}) or {}
    contract_label = contract.get("contract_name") or contract.get("contract_number") or "未识别合同名称"
    if contract.get("contract_name") and contract.get("contract_number"):
        contract_label = f"{contract['contract_name']}（{contract['contract_number']}）"
    duration = _duration_display(plan)
    delivery = _milestone_summary(plan, "到货")
    preliminary = _milestone_summary(plan, "初验")
    final = _milestone_summary(plan, "终验")
    return {
        "项目编码": project_code, "合同方向": direction, "合同名称/编号": contract_label,
        "签订日期": contract.get("sign_date") or NO_CONTRACT_RULE, "项目整体工期": duration,
        "开工起算方式": plan.get("start_condition_type", "没有明确"), "可计算开工日期": plan.get("start_date", ""),
        "可计算完工日期": plan.get("finish_date", ""),
        "到货时间要求": delivery[0], "到货计算日期": delivery[1],
        "初验时间要求": preliminary[0], "初验计算日期": preliminary[1],
        "终验时间要求": final[0], "终验计算日期": final[1],
    }


def _plan_review_row(project_code: str, difference: dict[str, Any]) -> dict[str, Any]:
    plan, contract = difference.get("forward", {}) or {}, difference.get("backward", {}) or {}
    plan_time = plan.get("财务计划日期", "")
    if plan.get("计划起始日期") or plan.get("计划截止日期"):
        plan_time = f"{plan.get('计划起始日期', '')}至{plan.get('计划截止日期', '')}"
    gaps = []
    if contract.get("偏差天数") is not None:
        gaps.append(f"{contract['偏差天数']:+d}天")
    if contract.get("开始偏差天数") is not None:
        gaps.append(f"开始{contract['开始偏差天数']:+d}天")
    if contract.get("截止偏差天数") is not None:
        gaps.append(f"截止{contract['截止偏差天数']:+d}天")
    source = "；".join(filter(None, (str(plan.get("来源文件") or ""), str(plan.get("来源工作表") or ""),
                                     str(plan.get("日期字段") or ""))))
    return {
        "项目编码": project_code, "计划类型": plan.get("计划类型") or difference.get("title", ""),
        "对应合同方向": contract.get("合同方向", ""),
        "对应合同": contract.get("合同名称") or contract.get("合同编号", ""),
        "复核节点": plan.get("复核节点") or difference.get("title", ""),
        "财务计划时间（基准）": plan_time, "合同解析时间": contract.get("合同解析日期", ""),
        "偏差": "；".join(gaps), "复核结论": difference.get("status", ""),
        "风险等级": difference.get("risk_level", ""), "复核说明": difference.get("description", ""),
        "计划来源": source,
    }


def _scope_summary(value: dict[str, Any]) -> str:
    if not value:
        return "未提取"
    original = str(value.get("original_text") or "").strip()
    if len(original) > 160:
        original = original[:157] + "..."
    return "；".join(filter(None, (str(value.get("responsibility") or ""), original)))


def _scope_review_row(project_code: str, difference: dict[str, Any]) -> dict[str, Any]:
    return {
        "项目编码": project_code, "主要实施环节": difference.get("title", ""),
        "前向要求": _scope_summary(difference.get("forward", {}) or {}),
        "后向安排": _scope_summary(difference.get("backward", {}) or {}),
        "比对结论": difference.get("status", ""), "风险等级": difference.get("risk_level", ""),
        "风险说明": difference.get("description", ""),
    }


def _safe_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .") or "未命名项目"


def export_review(store: ReviewStore, path: Path, project_codes: set[str] | None = None) -> Path:
    projects = [item["payload"] for item in store.list_projects()
                if project_codes is None or item["project_code"] in project_codes]
    summary: list[dict[str, Any]] = []
    contracts: list[dict[str, Any]] = []
    time_reviews: list[dict[str, Any]] = []
    equipment_rows: list[dict[str, Any]] = []
    scopes: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    for p in projects:
        equipment_differences = p.get("equipment_differences", [])
        summary.append({"项目编码": p["project_code"], "处理状态": p["status"], "风险等级": p["risk_level"],
                        "前向设备未覆盖数": len(equipment_differences),
                        "工期风险数": len(p.get("schedule_differences", [])),
                        "收入收款计划复核项数": len(p.get("plan_differences", [])),
                        "主要实施审查项数": len(p.get("scope_differences", [])), "待复核数": len(p.get("review_issues", [])),
                        "处理时间": p.get("processed_at", "")})
        if p.get("forward"):
            contracts.append(_contract_time_row(p["project_code"], "前向", p["forward"]))
        for index, contract in enumerate(p.get("backward_contracts", []), 1):
            contracts.append(_contract_time_row(p["project_code"], f"后向{index}", contract))
        for difference in equipment_differences:
            item = difference.get("forward", {})
            quantity = item.get("quantity")
            quantity_text = "" if quantity in (None, "") else f"{quantity:g}{item.get('unit', '')}"
            equipment_rows.append({
                "项目编码": p["project_code"],
                "前向设备名称": item.get("standard_name") or difference.get("title", ""),
                "前向品牌": item.get("brand", ""),
                "前向型号": item.get("model", ""),
                "前向数量": quantity_text,
                "后向查找结果": "未找到",
                "风险等级": difference.get("risk_level", "高风险"),
                "风险说明": difference.get("description", "后向合同未找到该前向设备"),
            })
        time_reviews.extend(_plan_review_row(p["project_code"], d) for d in p.get("plan_differences", []))
        scopes.extend(_scope_review_row(p["project_code"], d) for d in p.get("scope_differences", []))
        review_rows.extend({"记录类型": "待人工复核", "项目编码": p["project_code"], **x}
                           for x in p.get("review_issues", []))
    review_rows.extend({"记录类型": "人工纠正", **x} for x in store.list_corrections()
                       if project_codes is None or x.get("project_code") in project_codes)
    wb = Workbook(); wb.remove(wb.active)
    sheet_rows = [("项目审查汇总", summary), ("合同解析结果", contracts),
                  ("时间及收入计划复核", time_reviews), ("设备未覆盖风险", equipment_rows)]
    if scopes:
        sheet_rows.append(("实施内容差异", scopes))
    sheet_rows.append(("待复核及人工纠正", review_rows))
    for title, rows in sheet_rows:
        _add_sheet(wb, title, rows)
    path.parent.mkdir(parents=True, exist_ok=True); wb.save(path)
    return path


def export_project_reviews(store: ReviewStore, output_dir: Path, timestamp: str,
                           project_codes: set[str] | None = None) -> list[Path]:
    """Create one complete review workbook for every selected project."""
    paths: list[Path] = []
    for item in store.list_projects():
        project_code = item["project_code"]
        if project_codes is not None and project_code not in project_codes:
            continue
        paths.append(export_project_review(store, output_dir, project_code, timestamp))
    return paths


def export_project_review(store: ReviewStore, output_dir: Path, project_code: str, timestamp: str) -> Path:
    path = output_dir / f"{_safe_filename(project_code)}_前后向合同履约风险审查_{timestamp}.xlsx"
    return export_review(store, path, {project_code})
