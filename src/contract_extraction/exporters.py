from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ContractOutput
from .schema import EVIDENCE_FIELDS, RESULT_FIELDS, REVIEW_FIELDS


def save_checkpoint(output: ContractOutput, folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{output.result['合同号']}.json"
    payload = output.payload()
    payload["页面文本"] = [page.to_dict() for page in output.pages]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return path


def load_checkpoint(path: Path, fingerprint: str) -> dict[str, object] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return payload if payload.get("源文件指纹") == fingerprint else None


def _sheet(workbook: Workbook, title: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, 1):
        sample = [len(str(row.get(header, ""))) for row in rows[:100]]
        ws.column_dimensions[get_column_letter(index)].width = min(max([len(header) * 2, *sample, 10]), 45)
    if "待人工复核" in headers:
        review_col = headers.index("待人工复核") + 1
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, review_col).value == "是":
                for cell in ws[row]:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")


def export_excel(payloads: list[dict[str, object]], path: Path) -> Path:
    results = [item.get("结构化结果", {}) for item in payloads]
    evidences = [row for item in payloads for row in item.get("字段证据", [])]
    reviews = []
    for row in results:
        if row.get("待人工复核") == "是":
            reviews.append({"合同号": row.get("合同号", ""), "合同性质": row.get("合同性质", ""),
                            "主合同文件": row.get("主合同文件", ""), "复核原因": row.get("复核原因", ""),
                            "OCR质量": row.get("OCR质量", ""),
                            "建议复核动作": "打开字段证据定位页；签章日期优先核对600DPI增强图与原PDF。"})
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "合同提取结果", RESULT_FIELDS, results)
    _sheet(workbook, "字段证据", EVIDENCE_FIELDS, evidences)
    _sheet(workbook, "待人工复核", REVIEW_FIELDS, reviews)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
