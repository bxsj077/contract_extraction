from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .system_models import ContractStructured, Difference


SS = "urn:schemas-microsoft-com:office:spreadsheet"
ISO_DATE = re.compile(r"^20\d{2}-\d{1,2}-\d{1,2}$")


def _row_values(row: ET.Element) -> list[str]:
    values: list[str] = []
    column = 1
    for cell in row.findall(f"{{{SS}}}Cell"):
        index = cell.get(f"{{{SS}}}Index")
        if index:
            column = int(index)
        while len(values) < column - 1:
            values.append("")
        data = cell.find(f"{{{SS}}}Data")
        values.append("" if data is None or data.text is None else data.text.strip())
        merge = int(cell.get(f"{{{SS}}}MergeAcross", "0"))
        values.extend([""] * merge)
        column += 1 + merge
    while values and values[-1] == "":
        values.pop()
    return values


def _spreadsheetml(path: Path) -> dict[str, list[list[str]]]:
    root = ET.fromstring(path.read_text(encoding="utf-8").lstrip())
    result: dict[str, list[list[str]]] = {}
    for sheet in root.findall(f"{{{SS}}}Worksheet"):
        name = sheet.get(f"{{{SS}}}Name", "")
        table = sheet.find(f"{{{SS}}}Table")
        result[name] = [] if table is None else [_row_values(row) for row in table.findall(f"{{{SS}}}Row")]
    return result


def _xlsx(path: Path) -> dict[str, list[list[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {sheet.title: [["" if value is None else str(value).strip() for value in row]
                              for row in sheet.iter_rows(values_only=True)] for sheet in workbook.worksheets}
    finally:
        workbook.close()


def read_plan_workbook(path: Path) -> dict[str, list[list[str]]]:
    prefix = path.read_bytes()[:200].lstrip()
    if prefix.startswith(b"<?xml") and b"Workbook" in prefix:
        return _spreadsheetml(path)
    if path.suffix.lower() == ".xlsx" or prefix.startswith(b"PK"):
        return _xlsx(path)
    raise ValueError("暂不支持二进制旧版XLS，请另存为XLSX或Excel 2003 XML")


def _as_date(value: str) -> date | None:
    text = value.strip()[:10]
    if not ISO_DATE.fullmatch(text):
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _find_header(rows: list[list[str]], start_label: str, end_label: str) -> tuple[int, int, int] | None:
    for row_index, row in enumerate(rows):
        if start_label in row and end_label in row:
            return row_index, row.index(start_label), row.index(end_label)
    return None


def extract_plan_periods(path: Path) -> list[dict[str, Any]]:
    periods: list[dict[str, Any]] = []
    for sheet_name, rows in read_plan_workbook(path).items():
        definitions = []
        if "收入" in sheet_name:
            definitions.append(("收入计划", "起始日期", "终止日期"))
        if "收款" in sheet_name:
            definitions.append(("收款计划", "合同履约起始日期", "合同履约截止日期"))
        for plan_type, start_label, end_label in definitions:
            found = _find_header(rows, start_label, end_label)
            if not found:
                continue
            header_row, start_col, end_col = found
            for row_number, row in enumerate(rows[header_row + 1:], header_row + 2):
                if max(start_col, end_col) >= len(row):
                    continue
                start, end = _as_date(row[start_col]), _as_date(row[end_col])
                if start and end:
                    periods.append({"plan_type": plan_type, "start_date": start.isoformat(),
                                    "end_date": end.isoformat(), "file_name": path.name,
                                    "sheet_name": sheet_name, "row_number": row_number})
    return periods


def _combined_headers(rows: list[list[str]], marker: str) -> tuple[int, dict[str, int]] | None:
    for row_index, row in enumerate(rows):
        if marker not in row:
            continue
        combine_next = marker == "收款计划"
        next_row = rows[row_index + 1] if combine_next and row_index + 1 < len(rows) else []
        width = max(len(row), len(next_row))
        headers: dict[str, int] = {}
        for column in range(width):
            top = row[column].strip() if column < len(row) else ""
            bottom = next_row[column].strip() if column < len(next_row) else ""
            label = bottom or top
            if label:
                headers[label] = column
        return row_index + (1 if any(next_row) else 0), headers
    return None


def _row_value(row: list[str], headers: dict[str, int], *labels: str) -> str:
    for label in labels:
        column = headers.get(label)
        if column is not None and column < len(row) and row[column].strip():
            return row[column].strip()
    return ""


def _normalize_plan_node(text: str, fallback: str) -> str:
    compact = re.sub(r"\s+", "", text)
    mappings = (("到货", ("到货", "供货", "交货")), ("初验", ("初验", "进场验收")),
                ("终验", ("终验", "竣工验收", "最终验收")), ("开工", ("开工",)),
                ("试运行", ("试运行",)), ("上线", ("上线",)), ("项目完成", ("完工", "项目完成")))
    for node, keywords in mappings:
        if any(keyword in compact for keyword in keywords):
            return node
    return fallback


def extract_plan_nodes(path: Path) -> list[dict[str, Any]]:
    """Extract dated milestones from finance-reviewed income and collection plans."""
    nodes: list[dict[str, Any]] = []
    for sheet_name, rows in read_plan_workbook(path).items():
        if "收入" in sheet_name:
            found = _combined_headers(rows, "预计确认日期")
            if found:
                header_row, headers = found
                for row_number, row in enumerate(rows[header_row + 1:], header_row + 2):
                    actual = _row_value(row, headers, "实际确认日期")
                    expected = _row_value(row, headers, "预计确认日期")
                    value = actual or expected
                    parsed = _as_date(value)
                    if not parsed:
                        continue
                    context = " ".join(filter(None, (_row_value(row, headers, "里程碑节点"),
                                                       _row_value(row, headers, "摘要"),
                                                       _row_value(row, headers, "产品收入项"))))
                    nodes.append({"plan_type": "收入计划", "node": _normalize_plan_node(context, "收入确认"),
                                  "plan_date": parsed.isoformat(), "date_field": "实际确认日期" if actual else "预计确认日期",
                                  "source_text": context, "file_name": path.name, "sheet_name": sheet_name,
                                  "row_number": row_number})
        if "收款" in sheet_name:
            found = _combined_headers(rows, "收款计划")
            if found:
                header_row, headers = found
                for row_number, row in enumerate(rows[header_row + 1:], header_row + 2):
                    actual = _row_value(row, headers, "实际里程碑日期")
                    milestone = _row_value(row, headers, "预计里程碑日期")
                    collection = _row_value(row, headers, "预计收款日期")
                    value = actual or milestone or collection
                    parsed = _as_date(value)
                    if not parsed:
                        continue
                    context = " ".join(filter(None, (_row_value(row, headers, "里程碑节点"),
                                                       _row_value(row, headers, "履行要求"),
                                                       _row_value(row, headers, "第N次"))))
                    fallback = _row_value(row, headers, "履行要求") or "收款节点"
                    nodes.append({"plan_type": "收款计划", "node": _normalize_plan_node(context, fallback),
                                  "plan_date": parsed.isoformat(),
                                  "date_field": "实际里程碑日期" if actual else ("预计里程碑日期" if milestone else "预计收款日期"),
                                  "source_text": context, "file_name": path.name, "sheet_name": sheet_name,
                                  "row_number": row_number})
    return nodes


def _contract_node_date(contract: ContractStructured, node: str) -> str:
    plan = contract.time_plan
    if node == "开工":
        return plan.start_date or ""
    if node == "项目完成":
        return plan.finish_date or ""
    detail = plan.milestone_details.get(node, {})
    value = str(detail.get("计算日期") or plan.milestones.get(node, "") or "")[:10]
    return value if _as_date(value) else ""


def compare_income_collection_plan(contract: ContractStructured | None, files: list[str],
                                   tolerance_days: int = 31,
                                   backward_contracts: list[ContractStructured] | None = None) -> list[Difference]:
    if not files:
        return []
    periods_by_type: dict[str, list[dict[str, Any]]] = defaultdict(list)
    nodes: list[dict[str, Any]] = []
    errors: list[str] = []
    for file_name in files:
        try:
            plan_path = Path(file_name)
            for period in extract_plan_periods(plan_path):
                periods_by_type[period["plan_type"]].append(period)
            nodes.extend(extract_plan_nodes(plan_path))
        except Exception as exc:
            errors.append(f"{Path(file_name).name}：{exc}")
    results: list[Difference] = []
    if errors:
        results.append(Difference("收入收款计划", "计划文件解析失败", "待确认", "PL-000", "收入/收款计划",
                                  "；".join(errors), needs_review=True))
    if not periods_by_type and not nodes:
        results.append(Difference("收入收款计划", "未提取到时间节点", "待确认", "PL-001", "收入/收款计划",
                                  "计划文件存在，但未提取到预计确认日期、预计里程碑日期、预计收款日期或履约期间。",
                                  needs_review=True))
        return results
    for plan_type, periods in periods_by_type.items():
        plan_start = min(date.fromisoformat(x["start_date"]) for x in periods)
        plan_end = max(date.fromisoformat(x["end_date"]) for x in periods)
        source = {"计划类型": plan_type, "计划起始日期": plan_start.isoformat(), "计划截止日期": plan_end.isoformat(),
                  "计划行数": len(periods), "来源文件": "；".join(sorted({x["file_name"] for x in periods})),
                  "来源工作表": "；".join(sorted({x["sheet_name"] for x in periods}))}
        targets = ([contract] if plan_type == "收入计划" else (backward_contracts or []))
        if not targets:
            targets = [None]
        for target in targets:
            contract_data = {"合同方向": "前向" if plan_type == "收入计划" else "后向",
                             "合同名称": target.contract_name if target else "", "合同编号": target.contract_number if target else ""}
            if not target or not target.time_plan.start_date or not target.time_plan.finish_date:
                results.append(Difference("收入收款计划", "合同整体工期无法复核", "待确认", "PL-002", "项目整体工期",
                    f"以{plan_type}{plan_start}至{plan_end}为基准，但对应合同未提取到完整的开工和完工日期。",
                    source, contract_data, needs_review=True))
                continue
            contract_start, contract_end = date.fromisoformat(target.time_plan.start_date), date.fromisoformat(target.time_plan.finish_date)
            start_gap, end_gap = (contract_start - plan_start).days, (contract_end - plan_end).days
            contract_data.update({"合同解析日期": f"{contract_start}至{contract_end}",
                                  "开始偏差天数": start_gap, "截止偏差天数": end_gap})
            if start_gap == 0 and end_gap == 0:
                status, risk, review = "一致", "无风险", False
            elif max(abs(start_gap), abs(end_gap)) <= tolerance_days:
                status, risk, review = "基本一致", "无风险", False
            else:
                status, risk, review = "明显不一致", "中风险", True
            description = (f"以{plan_type}{plan_start}至{plan_end}为基准；合同解析为{contract_start}至{contract_end}；"
                           f"合同开始相对计划{start_gap:+d}天，截止相对计划{end_gap:+d}天。")
            results.append(Difference("收入收款计划", status, risk, "PL-003", "项目整体工期", description,
                                      source, contract_data,
                                      [f"{x['file_name']}#{x['sheet_name']}!{x['row_number']}" for x in periods], review))
    grouped_nodes: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for item in nodes:
        grouped_nodes[(item["plan_type"], item["node"], item["plan_date"])].append(item)
    for (plan_type, node, plan_date_text), items in grouped_nodes.items():
        plan_date = date.fromisoformat(plan_date_text)
        source = {"计划类型": plan_type, "复核节点": node, "财务计划日期": plan_date_text,
                  "日期字段": "；".join(sorted({x["date_field"] for x in items})), "计划行数": len(items),
                  "来源文件": "；".join(sorted({x["file_name"] for x in items})),
                  "来源工作表": "；".join(sorted({x["sheet_name"] for x in items}))}
        targets = ([contract] if plan_type == "收入计划" else (backward_contracts or [])) or [None]
        for target in targets:
            direction = "前向" if plan_type == "收入计划" else "后向"
            contract_data = {"合同方向": direction, "合同名称": target.contract_name if target else "",
                             "合同编号": target.contract_number if target else ""}
            parsed_date_text = _contract_node_date(target, node) if target else ""
            contract_data["合同解析日期"] = parsed_date_text
            if not parsed_date_text:
                status, risk, review = "合同节点缺失", "待确认", True
                description = f"以{plan_type}的{node}日期{plan_date_text}为基准，对应{direction}合同未提取到可比较日期。"
            else:
                gap = (date.fromisoformat(parsed_date_text) - plan_date).days
                contract_data["偏差天数"] = gap
                if gap == 0:
                    status, risk, review = "一致", "无风险", False
                elif abs(gap) <= tolerance_days:
                    status, risk, review = "基本一致", "无风险", False
                else:
                    status, risk, review = "明显不一致", "中风险", True
                description = (f"以{plan_type}的{node}日期{plan_date_text}为基准；合同解析日期{parsed_date_text}；"
                               f"合同相对计划{gap:+d}天。")
            results.append(Difference("收入收款计划", status, risk, "PL-NODE", node, description,
                                      source, contract_data,
                                      [f"{x['file_name']}#{x['sheet_name']}!{x['row_number']}" for x in items], review))
    return results
