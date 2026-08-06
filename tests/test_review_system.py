from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from contract_extraction.comparisons import compare_equipment, compare_schedule, compare_scopes
from contract_extraction.api import create_app
from contract_extraction.review_export import _header_cn, export_project_reviews, export_review
from contract_extraction.review_service import ReviewService
from contract_extraction.project_io import scan_projects
from contract_extraction.revenue_plan import compare_income_collection_plan, extract_plan_nodes, extract_plan_periods
from contract_extraction.structured import _extract_equipment, _extract_time_plan
from contract_extraction.models import PageText
from contract_extraction.storage import ReviewStore, finding_key
from contract_extraction.system_models import ContractStructured, Difference, EquipmentItem, ScopeItem, TimePlan


class ReviewSystemTests(unittest.TestCase):
    def contract(self, direction: str) -> ContractStructured:
        return ContractStructured("P001", direction, contract_type="集成实施类")

    def test_project_integrity_scan(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = Path(tmp) / "000123"; folder.mkdir()
            (folder / "前向合同.pdf").touch()
            project = scan_projects(Path(tmp))[0]
            self.assertEqual(project.project_code, "000123")
            self.assertEqual(project.status, "可解析单份合同")
            self.assertIn("缺少后向合同", project.issues)

    def test_api_uses_explicit_storage_paths(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "contracts"
            output = Path(tmp) / "results"
            app = create_app(root, output)
            config_route = next(route for route in app.routes if getattr(route, "path", "") == "/api/config")
            config = config_route.endpoint()
            self.assertEqual(config["合同上传根目录"], str(root))
            self.assertEqual(config["审查结果目录"], str(output))
            self.assertTrue(root.exists())
            self.assertTrue(any(getattr(route, "path", "") == "/api/tasks/{task_id}" for route in app.routes))
            self.assertTrue(any(getattr(route, "path", "") == "/api/tasks" for route in app.routes))
            self.assertTrue(any(getattr(route, "path", "") == "/api/projects/{project_code}" and
                                "DELETE" in getattr(route, "methods", set()) for route in app.routes))
            self.assertTrue(any(getattr(route, "path", "") ==
                                "/api/projects/{project_code}/findings/equipment/{finding_index}" and
                                "DELETE" in getattr(route, "methods", set()) for route in app.routes))
            self.assertTrue(any(getattr(route, "path", "") ==
                                "/api/projects/{project_code}/findings/{category}/{finding_index}" and
                                "PUT" in getattr(route, "methods", set()) for route in app.routes))
            fields_route = next(route.endpoint for route in app.routes
                                if getattr(route, "path", "") == "/api/correction-fields")
            fields = fields_route()
            self.assertIn({"field_path": "amount_yuan", "label": "合同金额（元）", "group": "合同基本信息"}, fields)
            self.assertTrue(any(item["field_path"] == "time_plan.milestone_details.到货.计算日期"
                                and item["group"] == "时间节点" for item in fields))

    def test_dismissed_equipment_finding_is_persisted_until_project_delete(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = ReviewStore(Path(tmp) / "review.db")
            finding = {
                "rule_id": "EQ-001", "title": "人脸识别终端设备",
                "forward": {"standard_name": "人脸识别终端设备", "brand": "海康威视",
                            "model": "DS-K1T673MW"},
            }
            expected = finding_key("equipment", finding)
            saved = store.dismiss_finding("P001", "equipment", finding, "人工确认后删除")
            self.assertEqual(saved["finding_key"], expected)
            self.assertEqual(store.dismissed_finding_keys("P001", "equipment"), {expected})
            store.delete_project("P001")
            self.assertEqual(store.dismissed_finding_keys("P001", "equipment"), set())

    def test_schedule_and_scope_finding_overrides_are_persisted(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = ReviewStore(Path(tmp) / "review.db")
            finding = {"rule_id": "SC-001", "title": "设备安装", "status": "实施内容缺失",
                       "risk_level": "高风险", "description": "自动判断", "forward": {}, "backward": {}}
            saved = store.save_finding_override(
                "P001", "scope", finding, "人工确认已覆盖", "无风险", "已在后向附件中确认", "人工复核")
            overrides = store.finding_overrides("P001", "scope")
            self.assertIn(saved["finding_key"], overrides)
            self.assertEqual(overrides[saved["finding_key"]]["risk_level"], "无风险")
            store.delete_project("P001")
            self.assertEqual(store.finding_overrides("P001", "scope"), {})

    def test_saved_finding_override_is_applied_to_recalculated_result(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            service = ReviewService(root / "contracts", root / "output")
            finding = Difference("实施内容", "实施内容缺失", "高风险", "SC-001", "设备安装", "自动判断")
            service.store.save_finding_override(
                "P001", "scope", finding, "人工确认已覆盖", "无风险", "已在后向附件中确认")
            applied = service._apply_finding_overrides("P001", "scope", [finding])[0]
            self.assertEqual((applied.status, applied.risk_level, applied.description),
                             ("人工确认已覆盖", "无风险", "已在后向附件中确认"))

    def test_home_supports_project_detail_and_section_hash_routes(self):
        with tempfile.TemporaryDirectory() as tmp:
            app = create_app(Path(tmp) / "contracts", Path(tmp) / "results")
            home = next(route.endpoint for route in app.routes if getattr(route, "path", "") == "/")
            html = home()
            self.assertIn('id="view-detail"', html)
            self.assertIn('id="detail-equipment"', html)
            self.assertIn('id="detailEditPanel"', html)
            self.assertIn('id="detailEditForm"', html)
            self.assertIn('id="findingEditForm"', html)
            self.assertIn('id="scopeRiskFilter"', html)
            self.assertIn("function openProject", html)
            self.assertIn("function openDetailEditor", html)
            self.assertIn("function openFindingEditor", html)
            self.assertIn("function renderScopeDifferences", html)
            self.assertIn("function deleteEquipmentFinding", html)
            self.assertIn("#project/${code}/equipment", html)
            self.assertIn("deleteProject(currentProjectCode)", html)

    def test_delete_project_removes_files_results_and_database_records(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "contracts"
            output = Path(tmp) / "results"
            app = create_app(root, output)
            service = app.state.service
            code = "DELETE001"
            project = root / code
            (project / "前向").mkdir(parents=True)
            (project / "后向").mkdir()
            (project / "前向" / "合同.pdf").touch()
            detail = output / "项目明细" / code
            detail.mkdir(parents=True)
            (detail / "结果.json").touch()
            cache = output / "ocr_cache" / f"{code}_前向"
            cache.mkdir(parents=True)
            service.store.upsert_project(code, str(project), "处理失败", "待确认", {"project_code": code})
            service.store.save_correction(code, "前向", "contract_name", "人工名称")
            endpoint = next(route.endpoint for route in app.routes
                            if getattr(route, "path", "") == "/api/projects/{project_code}"
                            and "DELETE" in getattr(route, "methods", set()))

            result = endpoint(code)

            self.assertEqual(result["status"], "项目已彻底删除")
            self.assertFalse(project.exists())
            self.assertFalse(detail.exists())
            self.assertFalse(cache.exists())
            self.assertIsNone(service.store.get_project(code))
            self.assertEqual(service.store.list_corrections(code), [])

    def test_single_project_root_with_multiple_backward_contracts(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "JSNJA2513970CGN00"
            (project / "前向").mkdir(parents=True)
            (project / "后向").mkdir()
            (project / "前向" / "主合同.pdf").touch()
            (project / "前向" / "附件.pdf").touch()
            (project / "后向" / "采购A.pdf").touch()
            (project / "后向" / "采购B.pdf").touch()
            found = scan_projects(project)[0]
            self.assertEqual(found.project_code, "JSNJA2513970CGN00")
            self.assertEqual(len(found.forward_pdfs), 2)
            self.assertEqual(len(found.backward_pdfs), 2)
            self.assertEqual(found.status, "可对比")

    def test_project_scan_detects_income_collection_plan(self):
        with tempfile.TemporaryDirectory() as tmp:
            project = Path(tmp) / "P100"
            (project / "前向").mkdir(parents=True)
            (project / "后向").mkdir()
            (project / "收入收款计划").mkdir()
            (project / "前向" / "前向.pdf").touch()
            (project / "后向" / "后向.pdf").touch()
            (project / "收入收款计划" / "计划.xls").touch()
            found = scan_projects(project)[0]
            self.assertEqual(len(found.revenue_plan_files), 1)

    def test_service_resource_table_is_extracted_as_service_objects(self):
        text = "\n".join(["资源内容", "资源类型", "数量", "含税单价", "高性能型-1", "40", "11,154.15",
                           "高性能型-2", "服务器资源", "425", "15,913.74",
                           "生产汇聚交换机资", "44", "34,419.10"])
        page = PageText("合同.pdf", "合同.pdf", 1, text, "OCR 300DPI", confidence=.95)
        evidence = []
        items = _extract_equipment("P1", "前向", [page], evidence)
        self.assertEqual([(x.standard_name, x.quantity) for x in items],
                         [("服务器资源服务-高性能型-1", 40), ("服务器资源服务-高性能型-2", 425),
                          ("生产汇聚交换机资源", 44)])
        self.assertTrue(all(x.list_type == "维保/服务对象清单" for x in items))

    def test_maintenance_resource_and_multi_page_spare_list_are_extracted(self):
        page1 = PageText("后向合同.pdf", "后向合同.pdf", 8, "\n".join([
            "高性能型-1（Redis）", "CPU：2*16核", "服务器", "维护1年", "40",
            "高性能型-2（应用虚拟化）", "服务器", "维护1年", "425",
            "高性能型-3（中间件服务器）", "服务器", "维护1年", "245",
            "高性能型-4（Mysql服务器）", "服务器", "维护1年", "35",
            "大二层控制器资源", "设备", "维护1年", "1", "原厂质保", "4",
            "生产汇聚交换机资源", "设备", "维护1年", "44",
            "管理网/备份网汇聚交换机资源", "设备", "维护1年", "29",
            "管理网/备份网接入交换机资源", "设备", "维护1年", "13", "合计", "91",
        ]), "原生文本", confidence=1.0)
        page2 = PageText("后向合同.pdf", "后向合同.pdf", 9, "\n".join([
            "六、现场备件库清单", "备件类型", "备件参数", "数量", "单位",
            "CPU", "Xeon Silver 4216@2.1GHz×2", "3", "块",
            "内存", "MEMORY H HMA84GR7DJR4N-WM", "12", "块",
        ]), "原生文本", confidence=1.0)
        page3 = PageText("后向合同.pdf", "后向合同.pdf", 10, "\n".join([
            "网卡", "10GB Ethernet converged network adapter", "1", "块",
            "硬盘", "SSD PM1643a/45a", "6", "块",
        ]), "原生文本", confidence=1.0)
        page4 = PageText("后向合同.pdf", "后向合同.pdf", 11, "\n".join([
            "生产汇聚交换机整机", "生产汇聚交换机整机", "1", "台",
            "40G 光模块", "40G 光模块", "5", "块",
            "乙方需建立起不少于上述清单描述的应急备件库",
        ]), "原生文本", confidence=1.0)
        items = _extract_equipment("P1", "后向", [page1, page2, page3, page4], [])
        resources = [item for item in items if item.list_type == "维保/服务对象清单"]
        spares = [item for item in items if item.list_type == "现场备件库清单"]
        self.assertEqual(len(resources), 8)
        self.assertEqual(next(item.quantity for item in resources if item.standard_name == "大二层控制器资源"), 5)
        self.assertEqual(len(spares), 6)
        self.assertEqual(spares[0].model, "Xeon Silver 4216@2.1GHz×2")
        self.assertEqual(spares[-1].standard_name, "40G 光模块")

    def test_cross_page_procurement_table_with_quantity_before_unit_is_extracted(self):
        header = "序号 名称 品牌 型号/软件版本号 数量 含税单价（单位：元） 含税合价（单位：元） 增值税税率"
        pages = [
            PageText("前向.pdf", "前向.pdf", 3, "\n".join([
                "南京卷烟厂", header,
                "1 监控摄像机 海康威视 DS-2CD264XV3-LD 73 台 525 38325 13%",
                "2 液位传感器 海康威视 NP-FSC210-4G 1 台 822 822 13%",
                "3 统一平台开发费用 海康威视 Infovision iWork-Safety 企业安全生产管理平台 1.9.101 1 套 182732 182732 13%",
            ]), "原生文本层", confidence=1.0),
            PageText("前向.pdf", "前向.pdf", 4, "\n".join([
                header,
                "3 热成像摄像机 海康威视 HM-TD26XS-D 2 台 2467 4934 13%",
                "淮阴卷烟厂", header,
                "1 手持巡查终端 康凯思特 conquest-F5 8 台 7081 56648 13%",
                "2 5G 布控球 海康威视 iDS-MCD432 2 套 23984 47968 13%",
            ]), "原生文本层", confidence=1.0),
        ]
        items = _extract_equipment("P1", "前向", pages, [])
        self.assertEqual(len(items), 6)
        self.assertEqual((items[0].standard_name, items[0].brand, items[0].model,
                          items[0].quantity, items[0].unit),
                         ("监控摄像机", "海康威视", "DS-2CD264XV3-LD", 73, "台"))
        self.assertEqual(items[-1].standard_name, "5G 布控球")
        self.assertEqual(items[-1].technical_parameters["清单分组"], "淮阴卷烟厂")
        self.assertTrue(all(item.list_type == "采购交付清单" for item in items))

    def test_product_table_continues_on_page_without_repeated_header(self):
        header = "序号 产品名称 品牌 规格型号 数量 单位 分项单价 分项总价 税率"
        pages = [
            PageText("前向.pdf", "前向.pdf", 7, "\n".join([
                "附件1.清单", header,
                "1 教学办公区智能管理平台 锐捷 RG-UNC-CS 1 套 338069 338069 13%",
                "2 教学办公区综合运维平台 锐捷 RG-Enjoy 1 套 500100 500100 13%",
            ]), "原生文本层", confidence=1.0),
            PageText("前向.pdf", "前向.pdf", 8, "\n".join([
                "3 万兆光汇聚交换机 锐捷 RG-S7610-10SFXP2CQ 3 台 212874 638622 13%",
                "4 集成服务费 中电鸿信 定制 1 项 1200000 1200000 6%",
                "以上合计（单位：元）7898800.00",
            ]), "原生文本层", confidence=1.0),
        ]
        items = _extract_equipment("P1", "前向", pages, [])
        self.assertEqual(len(items), 4)
        self.assertEqual(items[2].standard_name, "万兆光汇聚交换机")
        self.assertEqual((items[3].standard_name, items[3].category), ("集成服务费", "软件/服务"))

    def test_descriptive_procurement_table_is_joined_with_separate_price_table(self):
        pages = [
            PageText("后向.pdf", "后向.pdf", 26, "\n".join([
                "南京市（202603270017）信息化项目明细报价表",
                "序号 采购内容 技术参数 数量 单位",
                "1 监控设备1", "1.设备支持全景摄像头；", "详见技术规范书，", "8 台",
                "2 智能识别处理", "设备1", "支持危险作业监管、离岗睡岗；", "详见技术规范书，", "4 台",
            ]), "原生文本层", confidence=1.0),
            PageText("后向.pdf", "后向.pdf", 27, "\n".join([
                "3 水流量采集传", "感器（DN40）", "输出信号：4～20mA；", "详见技术规范书，", "2 台",
            ]), "原生文本层", confidence=1.0),
            PageText("后向.pdf", "后向.pdf", 30, "\n".join([
                "4 软件模块", "环境搭建、安装配置和系统联调；", "详见技术规范书，", "5 项",
                "设备部分合计（含税最高限价4766550元）",
                "1 需求调研与方", "案规划服务", "需求调研与方案规划，包含：", "1、业务需求梳理", "1 项",
                "服务部分合计（含税最高限价2100550元）",
            ]), "原生文本层", confidence=1.0),
            PageText("后向.pdf", "后向.pdf", 31, "\n".join([
                "品牌 型号 不含税单价（元） 增值税税率 含税单价（元） 不含税总价（元） 含税总价（元） 备注",
                "海康威视 iDS-", "MCD432 18580 13% 20995.4 148640 167963.2",
                "海康威视 iDS-", "96128NX-H16/HWF 44250 13% 50002.5 177000 200010",
                "海康威视 HM-FE00-F0040 3190 13% 3604.7 6380 7209.4",
            ]), "原生文本层", confidence=1.0),
            PageText("后向.pdf", "后向.pdf", 35, "\n".join([
                "/ 定制 99020 13% 111892.6 495100 559463",
                "设备部分合计（含税最高限价4766550元） 197170 474270.4 /",
                "/ / 429200 6% 454952 429200 454952",
                "服务部分合计（含税最高限价2100550元） 429200 454952 /",
                "总价（含税总限价6867100元) 6168780 6832702.4",
            ]), "原生文本层", confidence=1.0),
        ]
        evidence = []
        items = _extract_equipment("P1", "后向", pages, evidence)
        self.assertEqual(len(items), 5, [(item.standard_name, item.brand, item.model) for item in items])
        self.assertEqual((items[0].standard_name, items[0].brand, items[0].model,
                          items[0].quantity, items[0].unit),
                         ("监控设备1", "海康威视", "iDS-MCD432", 8, "台"))
        self.assertEqual(items[1].standard_name, "智能识别处理设备1")
        self.assertIn("危险作业监管", items[1].technical_parameters["技术参数"])
        self.assertEqual(items[2].standard_name, "水流量采集传感器（DN40）")
        self.assertEqual((items[3].standard_name, items[3].brand, items[3].model),
                         ("软件模块", "/", "定制"))
        self.assertEqual(items[4].standard_name, "需求调研与方案规划服务")
        self.assertEqual(items[4].category, "软件/服务")
        self.assertEqual(items[4].technical_parameters["增值税税率"], "6%")
        self.assertEqual(len(evidence), 5)

    def test_quantity_first_procurement_table_is_extracted_across_pages(self):
        header = "序号 名称 数量 单位 品牌 型号 增值税税率 含税单价（元） 含税总价（元）"
        pages = [
            PageText("后向.pdf", "后向.pdf", 2, "\n".join([
                "合同清单", header,
                "1 教学办公区智能管理平台 1 套 锐捷 RG-UNC-CS 13% 324546.24 324546.24",
                "2 教学办公区综合运维平台 1 套 锐捷 RG-Enjoy 13% 480096.00 480096",
                "3 万兆光汇聚交换机 3 台 锐捷 RG-S7610-10SFXP2CQ 13% 204359.04 613077.12",
                "4 2.5G 光汇聚交换 4 台 锐捷 RG-S7610-10SFXP2CQ 13% 192135.36 768541.44",
            ]), "原生文本层", confidence=1.0),
            PageText("后向.pdf", "后向.pdf", 3, "\n".join([
                header, "机",
                "5 48口万兆光汇聚 2 台 锐捷 RG-S6150-48VS8CQ-X 13% 188812.80 377625.6",
                "6 24口POE交换机 48 台 锐捷 RG-IF2920-24GT4MS-P 13% 2433.60 116812.8",
                "设备部分合计 6430848.00",
                "1 集成服务 1 项 南电 / 6% 1152000.00 1152000.00",
                "服务部分合计 1152000.00", "总计 7582848.00",
            ]), "原生文本层", confidence=1.0),
        ]
        items = _extract_equipment("P1", "后向", pages, [])
        self.assertEqual(len(items), 7)
        self.assertEqual((items[0].standard_name, items[0].brand, items[0].model,
                          items[0].quantity, items[0].unit),
                         ("教学办公区智能管理平台", "锐捷", "RG-UNC-CS", 1, "套"))
        self.assertEqual(items[3].standard_name, "2.5G光汇聚交换机")
        self.assertEqual(items[5].model, "RG-IF2920-24GT4MS-P")
        self.assertEqual((items[6].standard_name, items[6].category), ("集成服务", "软件/服务"))

    def test_compact_procurement_tables_with_and_without_brand_model_are_extracted(self):
        pages = [
            PageText("宏锦翔.pdf", "宏锦翔.pdf", 8, "\n".join([
                "清单：", "序号 采购内容 数量 单位 品牌 型号",
                "1 一体化云台监控设备 17 套 恩博 QRS7108-FMSZ3939",
                "2 智能识别报警设备 17 个 恩博 NB320-QRS111",
                "3 指挥中心大屏 1 套 创维", "SKYWORTH M55PJRGL-DS",
                "4 报警核实服务 26 路/3 年 / /", "]", "甲方：中电鸿信",
            ]), "原生文本层", confidence=1.0),
            PageText("铁塔.pdf", "铁塔.pdf", 8, "\n".join([
                "清单：", "序号 采购内容 数量 单位",
                "1 网络技术服务 1 18 条/3 年",
                "2 设备安装集成服务 2 17 项",
                "3 设备运营服务 17 项/3 年", "]", "甲方：中电鸿信",
            ]), "原生文本层", confidence=1.0),
        ]
        items = _extract_equipment("P1", "后向", pages, [])
        self.assertEqual(len(items), 7)
        self.assertEqual((items[0].standard_name, items[0].brand, items[0].model,
                          items[0].quantity, items[0].unit),
                         ("一体化云台监控设备", "恩博", "QRS7108-FMSZ3939", 17, "套"))
        self.assertEqual((items[2].standard_name, items[2].model),
                         ("指挥中心大屏", "SKYWORTH M55PJRGL-DS"))
        self.assertEqual((items[-2].standard_name, items[-2].quantity),
                         ("设备安装集成服务 2", 17))

    def test_hardware_supply_clause_is_selected_as_delivery_milestone(self):
        page = PageText("前向.pdf", "前向.pdf", 25,
                        "第一次付款：子合同签订后 40 日历日内完成硬\n件供货。"
                        "卖方应在交付前 7 日历日通知买方做好接收准备，到货经买方验收合格后付款。",
                        "原生文本层", confidence=1.0)
        plan = _extract_time_plan("P1", "前向", [page], {}, [])
        delivery = plan.milestone_details["到货"]
        self.assertIn("子合同签订后 40 日历日内完成硬件供货", delivery["原文"])
        self.assertIn("40 日历日内", delivery["相对期限"])
        self.assertEqual(delivery["计算状态"], "有明确相对期限，但缺少可确定的基准日期")

    def test_income_plan_period_comparison(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "计划.xls"
            path.write_text('''<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
<Worksheet ss:Name="收入计划"><Table>
<Row><Cell><Data ss:Type="String">起始日期</Data></Cell><Cell><Data ss:Type="String">终止日期</Data></Cell></Row>
<Row><Cell><Data ss:Type="String">2026-04-01</Data></Cell><Cell><Data ss:Type="String">2027-03-31</Data></Cell></Row>
</Table></Worksheet></Workbook>''', encoding="utf-8")
            periods = extract_plan_periods(path)
            self.assertEqual(periods[0]["start_date"], "2026-04-01")
            contract = self.contract("前向")
            contract.time_plan = TimePlan(12, "个月", "固定日期区间", start_date="2026-04-16", finish_date="2027-04-15")
            result = compare_income_collection_plan(contract, [str(path)])[0]
            self.assertEqual(result.status, "基本一致")
            self.assertIn("+15天", result.description)

    def test_income_and_collection_plan_milestones_are_extracted(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "计划.xls"
            path.write_text('''<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
<Worksheet ss:Name="收入计划"><Table>
<Row><Cell><Data ss:Type="String">里程碑节点</Data></Cell><Cell><Data ss:Type="String">预计确认日期</Data></Cell></Row>
<Row><Cell><Data ss:Type="String">设备到货</Data></Cell><Cell><Data ss:Type="String">2026-07-28</Data></Cell></Row>
</Table></Worksheet>
<Worksheet ss:Name="收款计划"><Table>
<Row><Cell><Data ss:Type="String">履行要求</Data></Cell><Cell><Data ss:Type="String">收款计划</Data></Cell></Row>
<Row><Cell ss:Index="2"><Data ss:Type="String">预计收款日期</Data></Cell></Row>
<Row><Cell><Data ss:Type="String">终验款</Data></Cell><Cell><Data ss:Type="String">2026-10-20</Data></Cell></Row>
</Table></Worksheet></Workbook>''', encoding="utf-8")
            nodes = extract_plan_nodes(path)
            self.assertEqual([(x["plan_type"], x["node"], x["plan_date"]) for x in nodes],
                             [("收入计划", "到货", "2026-07-28"), ("收款计划", "终验", "2026-10-20")])

    def test_export_headers_are_chinese(self):
        self.assertEqual(_header_cn("risk_level"), "风险等级")
        self.assertEqual(_header_cn("time_plan.start_date"), "实际起算日期")
        self.assertEqual(_header_cn("time_plan.milestones.终验"), "时间节点-终验")
        self.assertEqual(_header_cn("direction"), "结构化合同方向")

    def test_export_merges_duration_and_creates_one_workbook_per_project(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            store = ReviewStore(root / "review.db")
            payloads = [
                ("P001", {"duration_value": 12, "duration_unit": "个月", "duration_conclusion": "12个月"}),
                ("P002", {"duration_value": None, "duration_unit": "", "duration_conclusion": "按招标文件及投标文件执行"}),
            ]
            for code, plan in payloads:
                equipment_differences = ([{
                    "title": "数据库审计设备", "status": "后向未找到", "risk_level": "高风险",
                    "description": "后向合同未找到该前向设备",
                    "forward": {"standard_name": "数据库审计设备", "brand": "示例品牌",
                                "model": "DBA-1000", "quantity": 2.0, "unit": "台"},
                }] if code == "P001" else [])
                payload = {
                    "project_code": code, "status": "已完成", "risk_level": "低风险",
                    "forward": {"direction": "前向", "contract_number": code, "contract_name": f"{code}合同",
                                "time_plan": plan, "parse_metadata": {}},
                    "backward_contracts": [], "equipment_differences": equipment_differences, "schedule_differences": [],
                    "plan_differences": [], "scope_differences": [], "review_issues": [],
                }
                store.upsert_project(code, str(root / code), "已完成", "低风险", payload)

            full_path = export_review(store, root / "全量.xlsx")
            project_paths = export_project_reviews(store, root / "分项目审查结果", "20260805_120000")

            full_wb = load_workbook(full_path, read_only=True, data_only=True)
            rows = list(full_wb["合同解析结果"].values)
            sheet_names = full_wb.sheetnames
            equipment_rows = list(full_wb["设备未覆盖风险"].values)
            full_wb.close()
            self.assertIn("项目整体工期", rows[0])
            self.assertNotIn("工期数值", rows[0])
            self.assertNotIn("工期单位", rows[0])
            self.assertNotIn("工期关键判断", rows[0])
            self.assertNotIn("甲方", rows[0])
            self.assertNotIn("服务内容", rows[0])
            self.assertEqual(len(rows[0]), 14)
            duration_col = rows[0].index("项目整体工期")
            self.assertEqual([row[duration_col] for row in rows[1:]],
                             ["明确：12个月", "未明确：按招标文件及投标文件执行"])
            self.assertIn("设备未覆盖风险", sheet_names)
            self.assertEqual(equipment_rows[0], (
                "项目编码", "前向设备名称", "前向品牌", "前向型号", "前向数量",
                "后向查找结果", "风险等级", "风险说明"))
            self.assertEqual(equipment_rows[1][1:7],
                             ("数据库审计设备", "示例品牌", "DBA-1000", "2台", "未找到", "高风险"))
            self.assertEqual(len(project_paths), 2)
            for path in project_paths:
                wb = load_workbook(path, read_only=True, data_only=True)
                self.assertEqual(wb["项目审查汇总"].max_row, 2)
                self.assertEqual(wb["合同解析结果"].max_row, 2)
                wb.close()

    def test_equipment_with_matching_model_but_lower_quantity_is_listed(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.equipment = [EquipmentItem(standard_name="核心交换机", model="S6730", unit="台", quantity=2, evidence_id="F")]
        b.equipment = [EquipmentItem(standard_name="核心交换机", model="S6730", unit="台", quantity=1, evidence_id="B")]
        result = compare_equipment(f, b)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0].status, "后向数量不足")

    def test_equipment_model_and_quantity_match_tolerates_ocr_suffix(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.equipment = [EquipmentItem(standard_name="一体化云台摄像机恩博", model="QRS7108-FMSZ3939180",
                                     unit="套", quantity=17, evidence_id="F")]
        b.equipment = [EquipmentItem(standard_name="一体化云台监控设备", model="QRS7108-FMSZ3939",
                                     unit="套", quantity=17, evidence_id="B")]
        self.assertEqual(compare_equipment(f, b), [])

    def test_equipment_semantic_alias_and_quantity_match(self):
        f, b = self.contract("前向"), self.contract("后向汇总")
        f.equipment = [
            EquipmentItem(standard_name="设备挂载费国产优质项目配套 180", unit="项", quantity=16, evidence_id="F1"),
            EquipmentItem(standard_name="中间件东方通东方通 180", unit="套", quantity=2, evidence_id="F2"),
            EquipmentItem(standard_name="平山林场维护服务国产优质项目配套 180", unit="项", quantity=3, evidence_id="F3"),
        ]
        b.equipment = [
            EquipmentItem(standard_name="设备安装集成服务 1", unit="项", quantity=16, evidence_id="B1"),
            EquipmentItem(standard_name="中间模块服务", unit="套", quantity=2, evidence_id="B2"),
            EquipmentItem(standard_name="监控设备维护服务", unit="项", quantity=3, evidence_id="B3"),
        ]
        self.assertEqual(compare_equipment(f, b), [])

    def test_only_forward_equipment_not_found_backward_is_listed(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.equipment = [
            EquipmentItem(standard_name="核心交换机", model="S6730", evidence_id="F1"),
            EquipmentItem(standard_name="数据库审计设备", model="DBA-1000", evidence_id="F2"),
        ]
        b.equipment = [
            EquipmentItem(standard_name="交换机", model="S6730", evidence_id="B1"),
            EquipmentItem(standard_name="后向新增备件", model="SPARE-1", evidence_id="B2"),
        ]
        result = compare_equipment(f, b)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0].title, "数据库审计设备")
        self.assertEqual(result[0].status, "后向未找到")
        self.assertEqual(result[0].risk_level, "高风险")

    def test_schedule_late(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.time_plan = TimePlan(90, "日", "收到开工令", start_date="2026-01-01", finish_date="2026-04-01")
        b.time_plan = TimePlan(100, "日", "收到开工令", start_date="2026-01-01", finish_date="2026-04-11")
        self.assertEqual(compare_schedule(f, b, 15)[0].status, "明确来不及")

    def test_schedule_missing_never_says_satisfied(self):
        result = compare_schedule(self.contract("前向"), self.contract("后向"))[0]
        self.assertEqual(result.status, "缺少起算依据")
        self.assertTrue(result.needs_review)

    def test_responsibility_weakened(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.scopes = [ScopeItem("数据迁移", "负责完成", "全部历史数据", "全部", "", "前向", "", "F", .9)]
        b.scopes = [ScopeItem("数据迁移", "配合", "基础数据", "部分", "", "后向", "", "B", .9)]
        self.assertEqual(compare_scopes(f, b)[0].status, "责任程度弱化")

    def test_scope_comparison_excludes_supply_and_groups_installation_commissioning(self):
        f, b = self.contract("前向"), self.contract("后向")
        f.scopes = [
            ScopeItem("供货", "提供", "设备", "全部", "", "前向", "负责供货", "F1", .9),
            ScopeItem("安装", "负责完成", "设备", "全部", "", "前向", "负责安装", "F2", .9),
            ScopeItem("通电", "负责完成", "设备", "全部", "", "前向", "负责通电", "F3", .9),
        ]
        b.scopes = [ScopeItem("系统联调", "负责完成", "系统", "全部", "", "后向", "系统联调", "B1", .9)]
        result = compare_scopes(f, b)
        self.assertEqual([item.title for item in result], ["实施调试"])
        self.assertEqual(result[0].status, "完全覆盖")


if __name__ == "__main__":
    unittest.main()
